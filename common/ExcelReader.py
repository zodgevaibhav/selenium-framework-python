from openpyxl import load_workbook

class ExcelReader:

	def readData(filePath,sheetName):
		#wb=load_workbook(filename="../test/LoginTest.xlsx",read_only=True)
		wb=load_workbook(filename=filePath,read_only=True)
		#ws=wb["test_verifyLoginForInvalidUser"]
		ws=wb[sheetName]
		maxRow=ws.max_row
		maxColumn=ws.max_column
		print("**** Rows in sheet "+str(maxRow))
		print("**** Column in sheet "+str(maxColumn))
		data=[]
		for i in range(maxRow-1):
			tempRowArray=[]
			for j in range(maxColumn):
				#print(i,j)
				tempRowArray.append(ws.cell(i+2,j+1).value)
			data.append(tempRowArray)
		return data;